import frappe
import io
import os
import requests
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

@frappe.whitelist()
def export_excel_api(quotation_name):
    quotation = frappe.get_doc("Quotation", quotation_name)
    customer = frappe.get_doc("Customer", quotation.party_name)

    # Load template
    file_path = frappe.get_site_path("public", "files", "mẫu báo giá.xlsx")
    wb = load_workbook(file_path)
    ws = wb.active

    # Customer name
    ws["B9"] = customer.customer_name or ""

    # Get phone from Contact
    contact_name = frappe.db.get_value("Dynamic Link", {
        "link_doctype": "Customer",
        "link_name": customer.name,
        "parenttype": "Contact"
    }, "parent")

    contact_mobile = ""
    if contact_name:
        contact = frappe.get_doc("Contact", contact_name)
        contact_mobile = contact.mobile_no or contact.phone or ""
    ws["I9"] = contact_mobile

    # Get address from Address
    address_name = frappe.db.get_value("Dynamic Link", {
        "link_doctype": "Customer",
        "link_name": customer.name,
        "parenttype": "Address"
    }, "parent")

    address_display = ""
    if address_name:
        address = frappe.get_doc("Address", address_name)
        address_display = address.get("address_display") or ""
    ws["B10"] = address_display

    # Insert quotation items
    start_row = 14
    for i, item in enumerate(quotation.items):
        row = start_row + i
        ws[f"A{row}"] = i + 1

        # Merge B:C:D for item_name
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2).value = item.item_name

        ws[f"E{row}"] = item.description or ""
        ws[f"G{row}"] = item.item_code
        ws[f"H{row}"] = item.qty
        ws[f"L{row}"] = item.rate or 0
        ws[f"N{row}"] = item.amount or (item.qty * item.rate)

        # Insert image into I:J if available
        if item.image:
            try:
                image_path = ""
                if item.image.startswith("/files/"):
                    image_path = frappe.get_site_path("public", item.image.lstrip("/"))
                elif item.image.startswith("http"):
                    tmp_path = f"/tmp/tmp_item_{i}.png"
                    with open(tmp_path, "wb") as f:
                        f.write(requests.get(item.image).content)
                    image_path = tmp_path

                if os.path.exists(image_path):
                    img = XLImage(image_path)
                    img.width = 70
                    img.height = 70
                    ws.add_image(img, f"I{row}")
            except:
                pass  # Bỏ qua nếu có lỗi ảnh

    # Tổng cộng: Ghi vào cột N (column 14)
    ws.cell(row=17, column=14).value = quotation.total or 0
    ws.cell(row=18, column=14).value = 0
    ws.cell(row=19, column=14).value = 0
    ws.cell(row=20, column=14).value = quotation.total or 0

    # Xuất file về trình duyệt
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.local.response.filename = f"Bao_gia_{quotation.name}.xlsx"
    frappe.local.response.filecontent = output.read()
    frappe.local.response.type = "binary"
